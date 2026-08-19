# -*- coding: utf-8 -*-
"""
Created on Mon May 11 10:39:43 2026

@author: eloua
"""

"""aqt_output.py — Scrpit de lecture des sorties AQUATOX
lit le CSV de résultats et calcule les biomasses moyennes finales
par espèce, en ignorant le spin-up initial"""

import pandas as pd
import os
import re
import datetime

# Configuratioon
# ---------------------------------------------------------------------------------------

# Fraction de la simulation à ignorer (spin-up)
# exemple 0.5 = on ignore la première moitié, on moyenne sur la seconde moitié
SPINUP_FRACTION = 0.5

# Unités de biomasse dans les en-têtes CSV — on ne garde que ces colonnes
BIOMASS_UNITS = ["g/m2 dry", "mg/L dry"]

# Au-delà de cet écart (en jours) entre la date sélectionnée et la date
# anniversaire de VI, on avertit que le pas de sortie du CSV est trop
# grossier pour un alignement de date pertinent
DATE_MATCH_WARNING_DAYS = 30

# Époque des dates seriales Delphi (TDateTime) utilisées dans les .txt AQUATOX
_DELPHI_EPOCH = datetime.date(1899, 12, 30)

# lecture et traitement du csv
# -------------------------------------------------------------------------------

def read_results(csv_path, spinup_fraction=SPINUP_FRACTION, txt_path=None):
    """Lit le fichier CSV de résultats AQUATOX.
    Si txt_path est fourni, la date de VI (FirstDay dans le .txt) est
    utilisée pour ne moyenner, après le spin-up, que les points tombant
    à la même date calendaire chaque année: ça évite le biais de variations saisonnières 
    quand on moyenne tous les points post-spinup sans distinction.
    Si txt_path est omis, on retombe sur moyenne de tous les points post-spinup.
    Retourne un dict : nom_espèce -> biomasse_moyenne_finale (float)"""
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"CSV file not found : {csv_path}")

    print(f"Reading results : {os.path.basename(csv_path)}...")
    
    # Lecture csv
    df = pd.read_csv(csv_path, sep=",", skipinitialspace=True)

    # Nettoyage des noms de colonnes (espaces résiduels)
    df.columns = df.columns.str.strip()

    # Identification des colonnes de biomasse
    biomass_cols = _find_biomass_columns(df.columns)
    if not biomass_cols:
        raise ValueError("No biomass column found in CSV.")

    print(f"  {len(biomass_cols)} species detected in results.")

    # Sélection de la fenêtre post-spinup
    n_total = len(df)
    n_skip  = int(n_total * spinup_fraction)
    df_post = df.iloc[n_skip:].copy()

    # Sélection des points "même date que VI" pour éviter le biais saisonnier
    df_used = df_post
    if txt_path is not None and "Date" in df_post.columns:
        try:
            start_date = get_simulation_start_date(txt_path)
            selected = _select_matching_date_rows(df_post, start_date.month, start_date.day)
        except Exception as e:
            print(f" Date alignment impossible ({e}) -> back to classic mean.")
            selected = None

        if selected is not None and not selected.empty:
            max_diff = int(selected["_day_diff"].max())
            n_years  = selected["_year"].nunique()
            print(f"  {len(selected)} Point(s) retained: same date as VI "
                  f"({start_date.strftime('%d/%m')}), max gap {max_diff} j, "
                  f"in {n_years} year(s).")
            if max_diff > DATE_MATCH_WARNING_DAYS:
                print(f"  Important gap ({max_diff} j) : the CSV exit step "
                      f"might be to mutch to align on a reliable date.")
            df_used = selected
        elif selected is not None:
            print("  No usable dates after the spin-up -> return to the classic average.")
    elif txt_path is None:
        print("  (No txt_path provided -> classic average across all points post-spinup)")

    # Calcul des moyennes
    means = {}
    for col in biomass_cols:
        species_name = _clean_species_name(col)
        means[species_name] = float(df_used[col].mean())

    return means


# Alignement des dates (VI vs points post-spinup)
# -------------------------------------------------------------------------------

def _delphi_serial_to_date(serial):
    """Convertit une date seriale Delphi (TDateTime, jours depuis 1899-12-30)
    en date calendaire"""
    return _DELPHI_EPOCH + datetime.timedelta(days=int(float(serial)))


def get_simulation_start_date(txt_path):
    """Lit le fichier .txt AQUATOX et renvoie la date de début de simulation
    (champ "FirstDay" du SetupRecord): c'est la date à laquelle VI est fixée.$"""
    if not os.path.exists(txt_path):
        raise FileNotFoundError(f"Fichier .txt not found: {txt_path}")

    with open(txt_path, 'r', encoding='latin-1') as f:
        content = f.read()

    match = re.search(r'"FirstDay":\s*([\d.eE+-]+)', content)
    if not match:
        raise ValueError(f"Champ 'FirstDay' not found in {os.path.basename(txt_path)}")

    return _delphi_serial_to_date(match.group(1))


def _ref_day_of_year(month, day):
    """Jour dans l'année pour une année de référence non-bissextile
    (29 février ramené au 28 pour éviter les cas particuliers)"""
    if month == 2 and day == 29:
        day = 28
    return datetime.date(2001, month, day).timetuple().tm_yday


def _circular_day_diff(month1, day1, month2, day2):
    """Écart en jours entre deux dates calendaires (jour/mois), en tenant
    compte du bouclage autour du nouvel an (ex: 31/12 vs 02/01 -> 2 j, pas 363)"""
    d1 = _ref_day_of_year(month1, day1)
    d2 = _ref_day_of_year(month2, day2)
    diff = abs(d1 - d2)
    return min(diff, 365 - diff)


def _select_matching_date_rows(df_post, target_month, target_day, date_col="Date"):
    """Parmi les lignes post-spinup, garde pour chaque année la ligne dont la
    date est la plus proche de (target_month, target_day) la date de VI"""
    df_post = df_post.copy()
    df_post["_parsed_date"] = pd.to_datetime(df_post[date_col], dayfirst=True, errors="coerce")
    df_post = df_post.dropna(subset=["_parsed_date"])

    if df_post.empty:
        return df_post

    df_post["_year"] = df_post["_parsed_date"].dt.year
    df_post["_day_diff"] = df_post["_parsed_date"].apply(
        lambda d: _circular_day_diff(d.month, d.day, target_month, target_day)
    )

    idx_best = df_post.groupby("_year")["_day_diff"].idxmin()
    return df_post.loc[idx_best].sort_values("_year")

def read_final_values(csv_path, txt_path=None):
    """Retourne les biomasses finales du CSV AQUATOX.
    Retourne un dict : nom_espèce -> biomasse_finale (float)
    À utiliser pour le bilan VI vs V_finale (pas pour la calibration qui utilise
    la moyenne post-spinup sur les points alignés avec la date de VI).
    Si txt_path est fourni, "finale" veut dire : la dernière occurrence de la
    date de VI (jour/mois de FirstDay) disponible dans le CSV, plutôt que la
    toute dernière ligne — ça évite de comparer VI à un point pris au hasard
    dans le cycle saisonnier. Si txt_path est omis, comportement inchangé
    (dernière ligne du CSV)."""
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"CSV fimle not found: {csv_path}")

    df = pd.read_csv(csv_path, sep=",", skipinitialspace=True)
    df.columns = df.columns.str.strip()
    biomass_cols = _find_biomass_columns(df.columns)
    if not biomass_cols:
        raise ValueError("No biomass columns were found in the CSV")

    row_to_use = df.iloc[-1]

    if txt_path is not None and "Date" in df.columns:
        try:
            start_date = get_simulation_start_date(txt_path)
            selected = _select_matching_date_rows(df, start_date.month, start_date.day)
        except Exception as e:
            print(f" Date alignment impossible ({e}): final value = last line of CSV.")
            selected = None

        if selected is not None and not selected.empty:
            row_to_use = selected.iloc[-1]  # dernière année disponible à la date de VI
            diff = int(row_to_use["_day_diff"])
            print(f"  Final value  align on VI date ({start_date.strftime('%d/%m')}), "
                  f"gap {diff} j (année {int(row_to_use['_year'])}).")
            if diff > DATE_MATCH_WARNING_DAYS:
                print(f"  Important gap ({diff} j) : the CSV exit step "
                      f"is perhaps too crude for a reliable date alignment.")

    finals = {}
    for col in biomass_cols:
        species_name = _clean_species_name(col)
        finals[species_name] = float(row_to_use[col])

    return finals

def _find_biomass_columns(columns):
    """Identifie les colonnes contenant des biomasses
    On garde que les colonnes dont l'unité est dans BIOMASS_UNITS
    et dont le nom correspond à une espèce (pour éliminer les variable physique)"""
    biomass_cols = []
    for col in columns:
        for unit in BIOMASS_UNITS:
            if f"({unit})" in col:
                # Exclure les variables agrégées ou physiques
                if not any(excl in col for excl in [
                    "Benthic Invt", "Plankton Invt", "Fish Biomass",
                    "Nekton Invt", "Oyster", "Phyto. Biomass",
                    "GrowthRate2"
                ]):
                    biomass_cols.append(col)
                break
    return biomass_cols

def _clean_species_name(col_header):
    """ Extrait le nom de l'espèce depuis l'en-tête CSV.
    Ex: 'Tunas (g/m2 dry)' -> 'Tunas' """
    for unit in BIOMASS_UNITS:
        col_header = col_header.replace(f" ({unit})", "")
    return col_header.strip()

# Comparaison Vi / Vf
#--------------------------------------------------------------------------------------

def compare_to_initial(results_means, initial_conditions):
    """Compare les biomasses moyennes finales aux valeurs initiales observées.
    results_means      : dict nom_espèce → V_finale_moyenne (depuis read_results)
    initial_conditions : dict nom_espèce → VI (depuis parser.get_initial_conditions)
    Retourne un DataFrame avec les colonnes :
        Espèce | VI | V_finale | Ecart_abs | Ecart_pct | Converge"""
    rows = []

    for species, vi in initial_conditions.items():
        """ Correspondance entre le nom dans le .txt et le nom dans le CSV
         Les noms peuvent légèrement différer du coup on cherche une correspondance partielle"""
        v_finale = _find_matching_result(species, results_means)

        if v_finale is None:
            print(f"  WARNING : no '{species}' found in CSV results")
            continue

        ecart_abs = abs(v_finale - vi)
        ecart_pct = (ecart_abs / vi * 100) if vi > 0 else float('inf')
        converge  = ecart_pct < 20.0  # seuil : moins de 20% d'écart
        rows.append({
            "Espèce"    : species,
            "VI"        : vi,
            "V_finale"  : round(v_finale, 6),
            "Ecart_abs" : round(ecart_abs, 6),
            "Ecart_pct" : round(ecart_pct, 1),
            "Below threshold"  : "OUI" if converge else "NON",
        })

    df = pd.DataFrame(rows).sort_values("Ecart_pct", ascending=False)
    return df

def _find_matching_result(species_name, results_means):
    #Cherche le nom de l'espèce avec une hiérarchie dans la recherceh

    # 1. Correspondance exacte
    if species_name in results_means:
        return results_means[species_name]

    # 2. Extraction du nom entre crochets (ex: 'Diatoms1: [Phytoplankton]' -> 'Phytoplankton')
    match = re.search(r'\[([^\]]+)\]', species_name)
    if match:
        short_name = match.group(1)
        if short_name in results_means:
            return results_means[short_name]

    # 3. Correspondance partielle (seulement pour les noms longs > 3 caractères)
    # Pour éviter que 'pH' matche 'Phytoplankton'
    species_lower = species_name.lower()
    for key, val in results_means.items():
        key_lower = key.lower()
        if len(key_lower) > 3: # On ignore les clés trop courtes comme 'pH'
            if key_lower in species_lower or species_lower in key_lower:
                return val

    return None

#Affichage et export
#---------------------------------------------------------------------------------

def print_comparison(df_comparison):
    #Affiche un résumé dans la console
    n_total    = len(df_comparison)
    n_converge = (df_comparison["Converge"] == "OUI").sum()

    print(f"\n{'='*60}")
    print(f"  CALIBRATION report")
    print(f"{'='*60}")
    print(df_comparison.to_string(index=False))
    print(f"{'='*60}\n")


def export_comparison(df_comparison, output_path):
    #Exporte le tableau de comparaison en Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_comparison.to_excel(writer, index=False, sheet_name="Bilan_calibration")
        ws = writer.sheets["Bilan_calibration"]

        # Mise en forme : colonne "Below threshold" avec de la couleur
        from openpyxl.styles import PatternFill
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        status_col_name = "Below threshold" if "Below threshold" in df_comparison.columns else "Converge"
        converge_col = df_comparison.columns.get_loc(status_col_name) + 1
        for row_idx in range(2, len(df_comparison) + 2):
            cell = ws.cell(row=row_idx, column=converge_col)
            cell.fill = green if cell.value == "OUI" else red

        # Largeurs des colonnes pour un truc propre
        for col_idx, width in enumerate([30, 14, 14, 14, 12, 12], start=1):
            ws.column_dimensions[
                ws.cell(row=1, column=col_idx).column_letter
            ].width = width

    print(f"Bilan exporté : {os.path.basename(output_path)}")
